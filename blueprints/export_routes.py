from flask import Blueprint, send_file, redirect, url_for, flash, session, request, jsonify
from flask_login import login_required
import sqlite3
import pandas as pd
from io import BytesIO
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta

# Create logger
logger = logging.getLogger(__name__)

# Create blueprint
export_bp = Blueprint('export', __name__)

def view_required(f):
    """Decorator to check view permissions"""
    from functools import wraps
    from flask_login import current_user

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.has_view_permission():
            flash('You do not have permission to view inventory')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@export_bp.route('/export/<file_type>')
@view_required
def export_file(file_type):
    category = session.get('category', 'trailers')
    table_name = 'inventory' if category == 'trailers' else category

    try:
        conn = sqlite3.connect('inventory.db')
        df = pd.read_sql_query(f'SELECT * FROM {table_name}', conn)
        conn.close()

        # Remove ID column
        df = df.drop(columns=['id'], errors='ignore')

        output = BytesIO()
        if file_type == 'csv':
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(output, mimetype='text/csv', as_attachment=True, download_name=f'{category}.csv')
        elif file_type == 'xlsx':
            df.to_excel(output, index=False, engine='openpyxl')
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{category}.xlsx')
        else:
            flash('Invalid export format')
            return redirect(url_for('index'))
    except Exception as e:
        logger.error(f"Error exporting file: {e}")
        flash(f'Error exporting file: {str(e)}')
        return redirect(url_for('index'))

@export_bp.route('/export/facebook', methods=['GET', 'POST'])
@login_required
@view_required
def export_facebook():
    try:
        conn = sqlite3.connect('inventory.db')
        cursor = conn.cursor()

        # Check if this is a POST request with selected IDs
        if request.method == 'POST':
            data = request.json
            item_ids = data.get('ids', [])
            if not item_ids:
                return jsonify({'error': 'No items selected'}), 400
            
            placeholders = ','.join('?' * len(item_ids))
            cursor.execute(f'''
                SELECT * FROM inventory 
                WHERE id IN ({placeholders})
                AND sold = "No" 
                AND deleted_at IS NULL 
                ORDER BY id
            ''', item_ids)
        else:
            # Export everything (unsold only)
            cursor.execute('''
                SELECT * FROM inventory 
                WHERE sold = "No" 
                AND deleted_at IS NULL 
                ORDER BY id
            ''')
        
        items = cursor.fetchall()
        conn.close()

        # Prepare data for Facebook Marketplace format
        fb_data = []

        for item in items:
            # Map condition to Facebook's exact required values
            condition = item[8] or ''
            if condition.lower() in ['new', 'brand new']:
                fb_condition = 'New'
            elif condition.lower() in ['excellent', 'like new']:
                fb_condition = 'Used - Like New'
            elif condition.lower() in ['good', 'very good']:
                fb_condition = 'Used - Good'
            elif condition.lower() in ['fair', 'acceptable']:
                fb_condition = 'Used - Fair'
            else:
                fb_condition = 'Used - Good'

            # Build TITLE from Length, Year, Make, Type, VIN (max 150 chars)
            title_parts = []
            if item[1]:  # Length
                title_parts.append(f"{int(item[1])}FT")
            if item[2]:  # Year
                title_parts.append(str(item[2]))
            if item[3]:  # Make
                title_parts.append(item[3].upper())
            if item[4]:  # Type
                title_parts.append(item[4].upper())

            title = " ".join(title_parts)

            # Add full VIN to title if available
            if item[9]:  # VIN
                if title:
                    title += " - "
                title += item[9]

            # Truncate to 150 characters as required
            title = title[:150]

            # Build DESCRIPTION from Length x Width, Capacity, Description (max 5000 chars)
            desc_parts = []
            
            # Add Length x Width format (e.g., "12FT X 83IN")
            if item[1] and item[5]:  # Length and Width
                desc_parts.append(f"{int(item[1])}FT X {item[5]}IN")
            elif item[1]:  # Length only
                desc_parts.append(f"{int(item[1])}FT")
            elif item[5]:  # Width only
                desc_parts.append(f"{item[5]}IN")
            
            if item[6]:  # Capacity
                desc_parts.append(item[6].upper())

            description = ", ".join(desc_parts)

            # Add main description
            if item[7]:  # Description
                if description:
                    description += " - "
                description += item[7].upper()

            # Add LAST 7 of VIN to description - FIXED
            if item[9]:  # VIN
                vin_short = item[9][-7:] if len(item[9]) >= 7 else item[9]
                if description:
                    description += " "
                description += vin_short

            # Add location message
            if description:
                description += ". "
            description += "VISIT GOODNIGHT TRAILERS, LOCATED OFF I-27 IN CANYON!"

            # Truncate to 5000 characters as required
            description = description[:5000]

            # Get PRICE (required, as currency)
            # item[12] is sell_price (item[11] is hitch_type)
            price = float(item[12]) if item[12] else 0.0

            # Add row to Facebook data
            fb_data.append([title, price, fb_condition, description, 'Miscellaneous'])

        # Create Excel workbook manually to match exact template format
        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk Upload Template"

        # Row 1: Main title (merged across all columns)
        ws.merge_cells('A1:E1')
        cell_a1 = ws['A1']
        cell_a1.value = "Facebook Marketplace Bulk Upload Template"
        cell_a1.font = Font(bold=True, size=11)
        cell_a1.alignment = Alignment(horizontal='left', vertical='center')

        # Row 2: Instructions (merged across all columns)
        ws.merge_cells('A2:E2')
        cell_a2 = ws['A2']
        cell_a2.value = "You can create up to 50 listings at once. When you are finished, be sure to save or export this as an XLS/XLSX file."
        cell_a2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Row 3: Format requirements for each column
        format_requirements = [
            'REQUIRED | Plain text (up to 150 characters',
            'REQUIRED | A whole number in $',
            'REQUIRED | Supported values: "New"; "Used - Like New"; "Used - Good"; "Used - Fair"',
            'OPTIONAL | Plain text (up to 5000 characters)',
            'OPTIONAL | Type of listing'
        ]
        for col_idx, req in enumerate(format_requirements, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = req
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Row 4: Column headers
        headers = ['TITLE', 'PRICE', 'CONDITION', 'DESCRIPTION', 'CATEGORY']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        # Data rows starting from row 5
        for row_idx, row_data in enumerate(fb_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Format price column (column B) as currency
                if col_idx == 2:  # PRICE column
                    cell.number_format = '$#,##0.00'

        # Set column widths to match template
        ws.column_dimensions['A'].width = 50  # TITLE
        ws.column_dimensions['B'].width = 12  # PRICE
        ws.column_dimensions['C'].width = 20  # CONDITION
        ws.column_dimensions['D'].width = 80  # DESCRIPTION
        ws.column_dimensions['E'].width = 20  # CATEGORY

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='facebook_marketplace_export.xlsx'
        )

    except Exception as e:
        logger.error(f"Error exporting to Facebook format: {e}")
        flash(f'Error exporting to Facebook format: {str(e)}')
        return redirect(url_for('index'))
    
@export_bp.route('/export/squarespace')
@export_bp.route('/export/squarespace/<filter_type>')
@login_required
@view_required
def export_squarespace(filter_type='all'):
    try:
        conn = sqlite3.connect('inventory.db')
        cursor = conn.cursor()

        # Check if this is a POST request with selected IDs
        if request.method == 'POST':
            data = request.json
            item_ids = data.get('ids', [])
            if not item_ids:
                return jsonify({'error': 'No items selected'}), 400
            
            placeholders = ','.join('?' * len(item_ids))
            cursor.execute(f'''
                SELECT * FROM inventory 
                WHERE id IN ({placeholders})
                AND sold = "No" 
                AND deleted_at IS NULL 
                ORDER BY id
            ''', item_ids)
        else:
            # Export everything (unsold only)
            cursor.execute('''
                SELECT * FROM inventory 
                WHERE sold = "No" 
                AND deleted_at IS NULL 
                ORDER BY id
            ''')
        
        items = cursor.fetchall()
        conn.close()

        # Prepare data for Squarespace format (same as Facebook but without location and category)
        sq_data = []

        for item in items:
            # Map condition to standard values
            condition = item[8] or ''
            if condition.lower() in ['new', 'brand new']:
                sq_condition = 'New'
            elif condition.lower() in ['excellent', 'like new']:
                sq_condition = 'Used - Like New'
            elif condition.lower() in ['good', 'very good']:
                sq_condition = 'Used - Good'
            elif condition.lower() in ['fair', 'acceptable']:
                sq_condition = 'Used - Fair'
            else:
                sq_condition = 'Used - Good'

            # Build TITLE from Length, Year, Make, Type, VIN (max 150 chars)
            title_parts = []
            if item[1]:  # Length
                title_parts.append(f"{int(item[1])}FT")
            if item[2]:  # Year
                title_parts.append(str(item[2]))
            if item[3]:  # Make
                title_parts.append(item[3].upper())
            if item[4]:  # Type
                title_parts.append(item[4].upper())

            title = " ".join(title_parts)

            # Add full VIN to title if available
            if item[9]:  # VIN
                if title:
                    title += " - "
                title += item[9]

            # Truncate to 150 characters
            title = title[:150]

            # Build DESCRIPTION from Length x Width, Capacity, Description (NO LOCATION MESSAGE)
            desc_parts = []
            
                        # Add Year, Make, Type first
            if item[2]:  # Year
                desc_parts.append(str(item[2]))
            if item[3]:  # Make
                desc_parts.append(item[3].upper())
            if item[4]:  # Type
                desc_parts.append(item[4].upper())
                
            # Add Length x Width format
            if item[1] and item[5]:  # Length and Width
                desc_parts.append(f"{int(item[1])}FT X {item[5]}IN")
            elif item[1]:  # Length only
                desc_parts.append(f"{int(item[1])}FT")
            elif item[5]:  # Width only
                desc_parts.append(f"{item[5]}IN")
            
            if item[6]:  # Capacity
                desc_parts.append(item[6].upper())

            description = ", ".join(desc_parts)

            # Add main description
            if item[7]:  # Description
                if description:
                    description += " - "
                description += item[7].upper()

            # Add LAST 7 of VIN to description
            if item[9]:  # VIN
                vin_short = item[9][-7:] if len(item[9]) >= 7 else item[9]
                if description:
                    description += " "
                description += vin_short

            # Add PRICE to description
            price = float(item[12]) if item[12] else 0.0
            if price > 0:
                if description:
                    description += " - "
                description += f"${price:,.2f}"

            # NO LOCATION MESSAGE FOR SQUARESPACE

            # NO LOCATION MESSAGE FOR SQUARESPACE

            # Truncate to 5000 characters
            description = description[:5000]

            # Get PRICE
            price = float(item[12]) if item[12] else 0.0

            # Add row to Squarespace data (NO CATEGORY COLUMN)
            sq_data.append([title, price, sq_condition, description])

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Squarespace Export"

        # Row 1: Main title
        ws.merge_cells('A1:D1')
        cell_a1 = ws['A1']
        cell_a1.value = "Squarespace Marketplace Export"
        cell_a1.font = Font(bold=True, size=11)
        cell_a1.alignment = Alignment(horizontal='left', vertical='center')

        # Row 2: Instructions
        ws.merge_cells('A2:D2')
        cell_a2 = ws['A2']
        cell_a2.value = "Export for Squarespace or other marketplaces. Save as XLS/XLSX file."
        cell_a2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Row 3: Format requirements
        format_requirements = [
            'REQUIRED | Plain text (up to 150 characters)',
            'REQUIRED | A whole number in $',
            'REQUIRED | Condition',
            'OPTIONAL | Plain text (up to 5000 characters)'
        ]
        for col_idx, req in enumerate(format_requirements, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = req
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Row 4: Column headers
        headers = ['TITLE', 'PRICE', 'CONDITION', 'DESCRIPTION']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        # Data rows starting from row 5
        for row_idx, row_data in enumerate(sq_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Format price column
                if col_idx == 2:  # PRICE column
                    cell.number_format = '$#,##0.00'

        # Set column widths
        ws.column_dimensions['A'].width = 50  # TITLE
        ws.column_dimensions['B'].width = 12  # PRICE
        ws.column_dimensions['C'].width = 20  # CONDITION
        ws.column_dimensions['D'].width = 80  # DESCRIPTION

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='squarespace_export.xlsx'
        )

    except Exception as e:
        logger.error(f"Error exporting to Squarespace format: {e}")
        flash(f'Error exporting to Squarespace format: {str(e)}')
        return redirect(url_for('index'))